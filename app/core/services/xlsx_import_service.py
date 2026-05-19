"""Use case: importação de XLSX como tabelas Postgres no Raio X.

Fluxo:

  1. `preview(content, filename)` parseia o XLSX em memória (openpyxl), itera
     todas as abas, infere tipos por coluna a partir de até 1000 linhas de
     sample, e devolve uma estrutura editável + um `artifact_id` que armazena
     o arquivo bruto por 10 min no `artifact_store` (TTL).

  2. `import_sheets(artifact_id, sheet_specs, user)` lê o arquivo do
     artifact_store, aplica a decisão do usuário (skip/rename/type-override) e
     cria uma tabela por aba via `DynamicTableService.ensure_table()`, depois
     `bulk_insert()` as linhas. Persiste origem em `raiox_xlsx_origins`.

Decisões de design:

  - **Naming**: prefixo fixo `raiox_xlsx__<filename>__<sheet>` (todos sanitizados).
    O `SchemaService.list_tables(feature="raiox")` filtra por esse prefixo —
    tabelas xlsx aparecem só no Raio X, nunca no Radar/Churn/Admin.

  - **Type inference em camadas**: tenta BOOLEAN → BIGINT → DOUBLE PRECISION
    → DATE → TIMESTAMPTZ → TEXT. Adota o primeiro tipo onde ≥95% das células
    não-vazias parsearem com sucesso. Empty strings e None viram NULL e NÃO
    contam contra a porcentagem.

  - **Soft cap**: warnings em >10k linhas, >50 colunas, >25MB. Hard cap em
    100k linhas / 100 colunas / 50MB (levanta `ValueError`).

  - **`data_only=True`** no openpyxl: pega valor calculado das fórmulas
    (`=SUM(...)`), não o texto da fórmula. Arquivos abertos por Excel salvam
    o valor calculado em cache; arquivos gerados programaticamente podem
    devolver `None` — nesse caso o valor cai como NULL.

  - **Compartilhamento entre usuários do Raio X**: nome de tabela NÃO inclui
    username. Qualquer usuário com acesso ao Raio X vê todas as tabelas
    importadas. RBAC fica no nível do módulo (acesso a Raio X) — não na tabela.
"""

from __future__ import annotations

import io
import re
import unicodedata
import uuid
from datetime import date, datetime
from typing import Any

from app.adapters.db.postgres import connect, is_safe_ident, quote_ident
from app.adapters.db.repositories.raiox_repo import PgRaioXXlsxOriginRepository
from app.core.services.artifact_store import ArtifactStore
from app.core.services.dynamic_table_service import _AUDIT_COLS, DynamicTableService


# Prefixo fixo das tabelas importadas — `SchemaService` usa isso para
# filtrar por feature "raiox" sem precisar de tabela de meta.
TABLE_PREFIX = "raiox_xlsx__"

# Limites — soft (warning) vs hard (raise).
_SOFT_MAX_ROWS = 10_000
_SOFT_MAX_COLS = 50
_SOFT_MAX_BYTES = 25 * 1024 * 1024     # 25 MB
_HARD_MAX_ROWS = 100_000
_HARD_MAX_COLS = 100
_HARD_MAX_BYTES = 50 * 1024 * 1024     # 50 MB

# Type inference — uma coluna adota tipo X se >= 95% das células não-vazias
# parsearem com sucesso. Empty/None NÃO conta no denominador.
_TYPE_CONFIDENCE = 0.95
# Quantas linhas usar pra inferir tipo. Mais que isso vira diminishing returns
# e arquivos grandes ficam lentos.
_INFER_SAMPLE_SIZE = 1000

# Tipos Postgres aceitos como override do usuário. Outros viram TEXT por defesa.
_ALLOWED_PG_TYPES = {
    "TEXT", "BIGINT", "DOUBLE PRECISION", "DATE", "TIMESTAMPTZ", "BOOLEAN",
}


# ============================================================
# Sanitização e naming
# ============================================================


def _strip_accents(text: str) -> str:
    """Remove acentos para sanitização de nome de tabela. NFD + filtro Mn."""
    return "".join(
        c for c in unicodedata.normalize("NFD", text or "")
        if unicodedata.category(c) != "Mn"
    )


def _sanitize_for_pg(text: str, max_len: int = 30) -> str:
    """Converte string arbitrária em identificador Postgres seguro.

    Pipeline: strip-accents → lowercase → [a-z0-9_]+ → collapse → truncate.
    Vazio vira 'x'. Inicia com dígito vira `_<digit>...`.
    """
    text = _strip_accents(text or "").lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    if not text:
        text = "x"
    if text[0].isdigit():
        text = "_" + text
    return text[:max_len]


def build_table_name(filename: str, sheet_name: str) -> str:
    """`raiox_xlsx__<file>__<sheet>` (cada parte sanitizada + cap em 30)."""
    base = filename
    # remove extensão .xlsx se vier completa
    if base.lower().endswith(".xlsx"):
        base = base[:-5]
    elif base.lower().endswith(".xls"):
        base = base[:-4]
    file_part = _sanitize_for_pg(base, max_len=30)
    sheet_part = _sanitize_for_pg(sheet_name, max_len=25)
    name = f"{TABLE_PREFIX}{file_part}__{sheet_part}"
    # Postgres limite hard de 63 chars para identificadores
    return name[:63]


# ============================================================
# Type inference
# ============================================================


def _try_parse_bool(v: Any) -> bool | None:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        s = v.strip().lower()
        if s in {"true", "yes", "sim", "1", "verdadeiro"}:
            return True
        if s in {"false", "no", "nao", "não", "0", "falso"}:
            return False
    return None


def _try_parse_int(v: Any) -> int | None:
    if isinstance(v, bool):
        return None  # bool é subclass de int — não queremos colidir
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        # 1.0 conta como int; 1.5 não.
        if v.is_integer():
            return int(v)
        return None
    if isinstance(v, str):
        s = v.strip()
        # Estrito: int é só dígitos com sinal opcional. Strings com `,` ou `.`
        # NÃO contam como int — deixa o float parser tentar. Senão '1,5' viraria
        # 15 (removendo o ',') e quebraria a inference de coluna decimal BR.
        if re.fullmatch(r"[+-]?\d+", s):
            try:
                return int(s)
            except ValueError:
                return None
    return None


def _try_parse_float(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        # Aceita US (1,234.56) e BR (1.234,56) — heurística: se tem ',' e '.',
        # último símbolo é o decimal separator.
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _try_parse_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _try_parse_timestamp(v: Any) -> datetime | None:
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
            "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
        ):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def infer_column_type(values: list[Any]) -> str:
    """Retorna o tipo Postgres mais restritivo que cobre >=95% dos valores.

    Ordem (mais restritivo → mais permissivo):
      BOOLEAN → BIGINT → DOUBLE PRECISION → DATE → TIMESTAMPTZ → TEXT

    Empty/None são ignorados na contagem (não influenciam para nenhum lado).
    Lista 100% vazia vira TEXT (default safe).
    """
    non_empty = [v for v in values if not _is_empty(v)]
    if not non_empty:
        return "TEXT"
    total = len(non_empty)
    threshold = total * _TYPE_CONFIDENCE

    # Ordem importa: BOOLEAN antes de BIGINT porque 0/1 também parseia como int.
    parsers = [
        ("BOOLEAN", _try_parse_bool),
        ("BIGINT", _try_parse_int),
        ("DOUBLE PRECISION", _try_parse_float),
        ("DATE", _try_parse_date),
        ("TIMESTAMPTZ", _try_parse_timestamp),
    ]
    for type_name, parser in parsers:
        hits = sum(1 for v in non_empty if parser(v) is not None)
        if hits >= threshold:
            return type_name
    return "TEXT"


def coerce_value(value: Any, target_type: str) -> Any:
    """Coage `value` para o tipo Postgres alvo. Empty → None."""
    if _is_empty(value):
        return None
    if target_type == "BOOLEAN":
        return _try_parse_bool(value)
    if target_type == "BIGINT":
        return _try_parse_int(value)
    if target_type == "DOUBLE PRECISION":
        return _try_parse_float(value)
    if target_type == "DATE":
        return _try_parse_date(value)
    if target_type == "TIMESTAMPTZ":
        return _try_parse_timestamp(value)
    # TEXT (fallback): força str
    return str(value) if not isinstance(value, str) else value


# ============================================================
# Service
# ============================================================


class XlsxImportService:
    """Use case da importação. Stateless — toda I/O via artifact_store
    + repositórios injetados."""

    def __init__(
        self,
        artifact_store: ArtifactStore,
        origin_repo: PgRaioXXlsxOriginRepository | None = None,
        dynamic_table_service: DynamicTableService | None = None,
    ):
        self._artifacts = artifact_store
        self._origins = origin_repo or PgRaioXXlsxOriginRepository()
        self._dts = dynamic_table_service or DynamicTableService()

    # ---- Preview ----------------------------------------------------------

    async def preview(self, content: bytes, filename: str) -> dict:
        """Parseia o XLSX em memória, devolve preview + cacheia o arquivo
        para uso posterior por `import_sheets()`.

        `filename` é só o nome (sem path) — usado para naming e exibido na UI.
        """
        size = len(content)
        if size > _HARD_MAX_BYTES:
            raise ValueError(
                f"arquivo excede {_HARD_MAX_BYTES // 1024 // 1024}MB "
                f"(recebido {size // 1024 // 1024}MB)"
            )

        # `data_only=True` → openpyxl devolve o VALOR cacheado de fórmulas em
        # vez do texto. Se o arquivo nunca foi aberto no Excel, valor pode ser
        # None — toleramos como NULL.
        # `read_only=True` reduz uso de memória; útil para arquivos grandes.
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(content),
            data_only=True,
            read_only=True,
        )

        sheets_preview: list[dict] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets_preview.append(self._preview_sheet(filename, sheet_name, ws))

        wb.close()

        # Cacheia o arquivo bruto para o import poder reusar. TTL curto (artifact
        # store global tem 30min — suficiente; o usuário decide rápido).
        artifact = await self._artifacts.put(
            content=content,
            filename=filename,
            mime_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        return {
            "artifact_id": artifact.id,
            "original_filename": filename,
            "sheets": sheets_preview,
        }

    def _preview_sheet(
        self, filename: str, sheet_name: str, ws,
    ) -> dict:
        """Itera uma aba uma vez: pega headers, sample, infere tipos."""
        warnings: list[str] = []
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            # aba vazia — ainda preview, mas com 0 colunas
            return {
                "original_sheet_name": sheet_name,
                "suggested_table_name": build_table_name(filename, sheet_name),
                "columns": [],
                "rows_count": 0,
                "sample_rows": [],
                "warnings": ["aba vazia (sem header) — será ignorada se importada"],
            }

        # Normaliza header: vazio vira `col_<N>`, sanitiza nomes
        raw_headers = list(header_row)
        if len(raw_headers) > _HARD_MAX_COLS:
            raise ValueError(
                f"aba '{sheet_name}' excede {_HARD_MAX_COLS} colunas "
                f"(tem {len(raw_headers)})"
            )
        if len(raw_headers) > _SOFT_MAX_COLS:
            warnings.append(
                f"{len(raw_headers)} colunas (acima de {_SOFT_MAX_COLS} recomendado) — "
                "performance de queries pode degradar"
            )

        # Colide se 2 headers sanitizam para o mesmo nome — desambigua com sufixo.
        seen: dict[str, int] = {}
        original_headers: list[str] = []
        sanitized_headers: list[str] = []
        for idx, h in enumerate(raw_headers):
            original = str(h).strip() if h is not None else f"col_{idx + 1}"
            if not original:
                original = f"col_{idx + 1}"
            sanitized = _sanitize_for_pg(original, max_len=40)
            # Evita choque com auditoria do DynamicTableService
            if sanitized in _AUDIT_COLS:
                sanitized = f"{sanitized}_x"
            base = sanitized
            n = seen.get(base, 0)
            if n > 0:
                sanitized = f"{base}_{n + 1}"
            seen[base] = n + 1
            original_headers.append(original)
            sanitized_headers.append(sanitized)

        # Itera o resto coletando sample para inference + count
        n_cols = len(sanitized_headers)
        col_samples: list[list[Any]] = [[] for _ in range(n_cols)]
        sample_rows: list[dict[str, Any]] = []  # primeiros 5
        rows_count = 0

        for row in rows_iter:
            rows_count += 1
            if rows_count > _HARD_MAX_ROWS:
                raise ValueError(
                    f"aba '{sheet_name}' excede {_HARD_MAX_ROWS} linhas — "
                    "filtre os dados no Excel antes de importar"
                )
            for i, cell in enumerate(row[:n_cols]):
                if i < n_cols and len(col_samples[i]) < _INFER_SAMPLE_SIZE:
                    col_samples[i].append(cell)
            if len(sample_rows) < 5:
                sample_rows.append({
                    sanitized_headers[i]: _to_jsonable(row[i] if i < len(row) else None)
                    for i in range(n_cols)
                })

        if rows_count > _SOFT_MAX_ROWS:
            warnings.append(
                f"{rows_count} linhas (acima de {_SOFT_MAX_ROWS} recomendado) — "
                "considere filtrar no Excel"
            )

        columns: list[dict] = []
        for i, sanitized in enumerate(sanitized_headers):
            inferred = infer_column_type(col_samples[i])
            samples_str = [
                str(_to_jsonable(v))[:80]
                for v in col_samples[i][:5] if not _is_empty(v)
            ][:3]
            columns.append({
                "original_name": original_headers[i],
                "sanitized_name": sanitized,
                "inferred_type": inferred,
                "sample_values": samples_str,
            })

        return {
            "original_sheet_name": sheet_name,
            "suggested_table_name": build_table_name(filename, sheet_name),
            "columns": columns,
            "rows_count": rows_count,
            "sample_rows": sample_rows,
            "warnings": warnings,
        }

    # ---- Import -----------------------------------------------------------

    async def import_sheets(
        self,
        artifact_id: str,
        sheet_specs: list[dict],
        user_id: str | None,
        username: str,
    ) -> list[dict]:
        """Cria as tabelas (uma por aba não-skipada). Idempotente: se a tabela
        já existe, faz `DROP TABLE` antes (reimport substitui)."""
        artifact = await self._artifacts.get(artifact_id)
        if not artifact:
            raise ValueError(
                "artifact expirado ou inválido — refaça o upload (TTL 30min)"
            )

        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(artifact.content),
            data_only=True,
            read_only=True,
        )

        # Index spec por nome original da aba pra lookup O(1)
        by_sheet = {s["original_sheet_name"]: s for s in sheet_specs}

        results: list[dict] = []
        for sheet_name in wb.sheetnames:
            spec = by_sheet.get(sheet_name)
            if not spec:
                continue  # aba não foi mencionada na request → ignora
            if spec.get("skip", False):
                results.append({
                    "table_name": spec.get("table_name", ""),
                    "rows_inserted": 0,
                    "columns_count": 0,
                    "status": "skipped",
                    "error": "",
                })
                continue

            ws = wb[sheet_name]
            try:
                rows_inserted, n_cols = await self._import_sheet(
                    ws=ws,
                    spec=spec,
                    filename=artifact.filename,
                    user_id=user_id,
                    username=username,
                )
                results.append({
                    "table_name": spec["table_name"],
                    "rows_inserted": rows_inserted,
                    "columns_count": n_cols,
                    "status": "created",
                    "error": "",
                })
            except Exception as exc:  # noqa: BLE001
                results.append({
                    "table_name": spec.get("table_name", ""),
                    "rows_inserted": 0,
                    "columns_count": 0,
                    "status": "failed",
                    "error": str(exc)[:300],
                })

        wb.close()
        return results

    async def _import_sheet(
        self,
        *,
        ws,
        spec: dict,
        filename: str,
        user_id: str | None,
        username: str,
    ) -> tuple[int, int]:
        """Cria a tabela física + bulk insert. Devolve (rows_inserted, n_cols)."""
        # 1. Re-sanitiza o nome (o usuário pode ter editado livremente no modal)
        table_name = spec["table_name"]
        # Garante prefixo `raiox_xlsx__` — nunca permitir o usuário escapar disso,
        # senão a tabela vaza para outras features no SchemaService.
        if not table_name.startswith(TABLE_PREFIX):
            table_name = TABLE_PREFIX + _sanitize_for_pg(table_name, max_len=50)
        table_name = _sanitize_for_pg(table_name, max_len=63)
        if not is_safe_ident(table_name):
            raise ValueError(f"nome de tabela inválido: {table_name}")

        # 2. Re-parseia cabeçalho desta aba
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return 0, 0

        # Reconstrói sanitized_headers do mesmo jeito do preview pra garantir
        # consistência (o user pode ter visto outro nome no modal por bug).
        raw_headers = list(header_row)
        seen: dict[str, int] = {}
        sanitized_headers: list[str] = []
        for idx, h in enumerate(raw_headers):
            original = str(h).strip() if h is not None else f"col_{idx + 1}"
            if not original:
                original = f"col_{idx + 1}"
            sanitized = _sanitize_for_pg(original, max_len=40)
            if sanitized in _AUDIT_COLS:
                sanitized = f"{sanitized}_x"
            base = sanitized
            n = seen.get(base, 0)
            if n > 0:
                sanitized = f"{base}_{n + 1}"
            seen[base] = n + 1
            sanitized_headers.append(sanitized)

        n_cols = len(sanitized_headers)
        if n_cols == 0:
            return 0, 0

        # 3. Decide tipo final por coluna: override do usuário ou inferência
        overrides = {
            c["sanitized_name"]: c["type"].upper().strip()
            for c in spec.get("column_types_override", [])
            if c.get("type", "").upper().strip() in _ALLOWED_PG_TYPES
        }
        # Coleta sample novamente para inferência (não evitamos o trabalho — barato).
        # Otimização possível futura: o caller poderia mandar os tipos no spec.
        col_samples: list[list[Any]] = [[] for _ in range(n_cols)]
        for row in ws.iter_rows(min_row=2, max_row=_INFER_SAMPLE_SIZE + 1, values_only=True):
            for i, cell in enumerate(row[:n_cols]):
                col_samples[i].append(cell)
        col_types: list[str] = []
        for i, sanitized in enumerate(sanitized_headers):
            t = overrides.get(sanitized) or infer_column_type(col_samples[i])
            if t not in _ALLOWED_PG_TYPES:
                t = "TEXT"
            col_types.append(t)

        # 4. Drop + recria (reimport = substitui)
        async with connect() as db:
            await db.execute(f"DROP TABLE IF EXISTS {quote_ident(table_name)}")

        # 5. Cria a tabela com tipos *desejados* (não delegamos pro DynamicTableService
        #    porque ele infere apenas a partir de uma única row e força BIGINT/TEXT).
        cols_sql = [f"{quote_ident(c)} {t}" for c, t in _AUDIT_COLS.items()]
        for sanitized, t in zip(sanitized_headers, col_types):
            if sanitized in _AUDIT_COLS:
                continue
            cols_sql.append(f"{quote_ident(sanitized)} {t}")
        ddl = (
            f"CREATE TABLE {quote_ident(table_name)} ("
            + ", ".join(cols_sql)
            + ")"
        )
        async with connect() as db:
            await db.execute(ddl)
            await db.execute(
                f"CREATE INDEX IF NOT EXISTS "
                f"{quote_ident('idx_' + table_name + '_ts')} "
                f"ON {quote_ident(table_name)}(_ts DESC)"
            )

        # 6. Bulk insert (re-itera do começo — read_only do openpyxl é uma vez só).
        rows_inserted = 0
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter)  # pula header
        BATCH = 500
        batch_rows: list[list[Any]] = []
        non_audit_cols = [c for c in sanitized_headers if c not in _AUDIT_COLS]
        non_audit_types = [
            t for c, t in zip(sanitized_headers, col_types) if c not in _AUDIT_COLS
        ]
        all_cols = list(_AUDIT_COLS.keys()) + non_audit_cols
        n_total_cols = len(all_cols)

        async def flush(batch: list[list[Any]]) -> int:
            if not batch:
                return 0
            placeholders_per_row = ", ".join(f"${i + 1}" for i in range(n_total_cols))
            # asyncpg quer rows posicionais — usamos executemany.
            async with connect() as db:
                await db.executemany(
                    f"INSERT INTO {quote_ident(table_name)} "
                    f"({', '.join(quote_ident(c) for c in all_cols)}) "
                    f"VALUES ({placeholders_per_row})",
                    batch,
                )
            return len(batch)

        for row in rows_iter:
            if rows_inserted + len(batch_rows) >= _HARD_MAX_ROWS:
                break
            audit_values = [
                uuid.uuid4().hex,  # _id
                None,              # _ts → uses DEFAULT NOW() via NULL? NO — explicit NOW.
                user_id,           # _user_id
                username,          # _username
                None,              # _case_number
                None,              # _transaction_id
                "raiox_xlsx",      # _feature
            ]
            # NULL para _ts faz o PG ignorar e usar DEFAULT? Não — INSERT VALUES com NULL
            # explícito sobrescreve o default. Para usar default, omitiríamos a coluna.
            # Mais simples: pegar `datetime.utcnow()` agora.
            audit_values[1] = datetime.utcnow()

            data_values = []
            for i, sanitized in enumerate(sanitized_headers):
                if sanitized in _AUDIT_COLS:
                    continue
                raw = row[i] if i < len(row) else None
                target = non_audit_types[non_audit_cols.index(sanitized)]
                data_values.append(coerce_value(raw, target))

            batch_rows.append(audit_values + data_values)
            if len(batch_rows) >= BATCH:
                rows_inserted += await flush(batch_rows)
                batch_rows = []

        rows_inserted += await flush(batch_rows)

        # 7. Registra origem
        await self._origins.upsert(
            table_name=table_name,
            original_filename=filename,
            sheet_name=spec["original_sheet_name"],
            uploaded_by_id=user_id,
            uploaded_by_username=username,
            rows_count=rows_inserted,
            columns_count=len(non_audit_cols),
        )

        return rows_inserted, len(non_audit_cols)

    # ---- Delete -----------------------------------------------------------

    async def delete_table(self, table_name: str) -> bool:
        """Drop atômico da tabela + remoção da origem.

        Só aceita nomes com o prefixo `raiox_xlsx__` — garante que o endpoint
        não vire ferramenta de SQL injection para apagar `users` ou `modules`.
        """
        if not table_name.startswith(TABLE_PREFIX):
            raise ValueError(
                f"só é possível deletar tabelas com prefixo {TABLE_PREFIX!r}"
            )
        if not is_safe_ident(table_name):
            raise ValueError(f"nome de tabela inválido: {table_name}")
        async with connect() as db:
            await db.execute(f"DROP TABLE IF EXISTS {quote_ident(table_name)}")
        return await self._origins.delete(table_name)


# ============================================================
# Helpers
# ============================================================


def _to_jsonable(v: Any) -> Any:
    """Converte tipos não-JSON (date, datetime, etc.) para algo serializável."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    return str(v)
