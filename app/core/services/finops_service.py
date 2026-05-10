"""Use case: FinOps Ledger, Orçamentos, Políticas e Cost-Aware Routing."""

from __future__ import annotations

import io
import json as _json
from dataclasses import dataclass
from typing import Any
from uuid import UUID

from app.core.domain.entities import (
    BudgetPeriod,
    BudgetScopeType,
    FinOpsAlert,
    FinOpsBudget,
    FinOpsEntry,
    FinOpsModelPolicy,
    RiskTier,
    ValueTier,
    new_uuid,
)
from app.core.ports.repositories import (
    FinOpsBudgetRepository,
    FinOpsModelPolicyRepository,
    FinOpsRepository,
)


# =============================================================================
# XLSX templates (download + import)
#
# Cada template tem uma única aba com header explícito + 2 linhas de exemplo.
# A ingestão é tolerante: mapeia case-insensitive, ignora linhas vazias e
# retorna lista de erros por linha em vez de abortar tudo.
# =============================================================================


_BUDGET_HEADERS = [
    "name", "scope_type", "scope_value", "period",
    "limit_brl", "warning_threshold", "hard_stop", "notes",
]

_BUDGET_EXAMPLES = [
    ["Radar mensal", "module", "radar", "monthly", 500.00, 0.8, "false", "limite operacional radar"],
    ["GPT-4o global", "model", "gpt-4o", "monthly", 1000.00, 0.9, "true", "hard stop ativo"],
]

_POLICY_HEADERS = [
    "model_name", "risk_tier", "value_tier", "max_cost_per_call",
    "max_tokens_per_call", "allowed_features", "rationale", "enabled",
]

_POLICY_EXAMPLES = [
    ["gpt-4o", "medium", "high", 0.50, 4096, "radar,churn", "modelo premium para análises", "true"],
    ["sabia-4", "low", "medium", 0.05, 2048, "", "modelo padrão pt-BR", "true"],
]


def _build_xlsx_template(
    sheet_name: str,
    headers: list[str],
    examples: list[list[Any]],
    helper: str,
) -> bytes:
    """Gera bytes de um xlsx com header em negrito + exemplos."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # linha 1: helper
    ws.cell(row=1, column=1, value=helper).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # linha 2: header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")

    # linhas 3+: exemplos
    for r_off, row in enumerate(examples, 3):
        for c_off, val in enumerate(row, 1):
            ws.cell(row=r_off, column=c_off, value=val)

    # auto-ajusta largura das colunas (heurística simples)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = max(14, len(h) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_xlsx_rows(file_bytes: bytes, expected_headers: list[str]) -> list[dict]:
    """Lê o xlsx e devolve lista de dicts (já filtrando linhas vazias e
    a linha de helper que pode aparecer antes do header)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Localiza linha do header — primeira linha cuja primeira célula bate com
    # qualquer expected_header. Tolera 0..N linhas de helper antes.
    expected_lower = {h.lower() for h in expected_headers}
    header_row_idx = None
    for i, row in enumerate(rows):
        first = row[0]
        if first and str(first).strip().lower() in expected_lower:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError(
            f"cabeçalho não encontrado. Esperado pelo menos uma destas colunas "
            f"como primeira célula: {expected_headers}"
        )

    header = [str(h).strip().lower() if h else "" for h in rows[header_row_idx]]
    idx_map = {h: i for i, h in enumerate(header) if h in expected_lower}

    out: list[dict] = []
    for row in rows[header_row_idx + 1:]:
        if not any(c not in (None, "") for c in row):
            continue
        rec: dict[str, Any] = {}
        for h, i in idx_map.items():
            v = row[i] if i < len(row) else None
            if isinstance(v, str):
                v = v.strip() or None
            rec[h] = v
        out.append(rec)
    return out


def _coerce_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"true", "1", "yes", "sim", "y", "s"}


class FinOpsService:
    def __init__(self, repo: FinOpsRepository):
        self.repo = repo

    async def record(
        self,
        user_id,
        module_id,
        model_name: str,
        tokens_input: int,
        tokens_output: int,
        cost_estimated: float,
        context_tag: str = "",
        domain: str | None = None,
        product: str | None = None,
        agent: str | None = None,
        flow: str | None = None,
        prompt_id: str | None = None,
        integration: str | None = None,
        environment: str = "production",
        latency_ms: float | None = None,
        storage_bytes: int | None = None,
    ) -> FinOpsEntry:
        """Registra uma chamada no ledger.

        Os campos extras são opcionais — callers antigos continuam válidos.
        Encorajamos preencher pelo menos ``agent``, ``flow`` e ``prompt_id``
        quando disponíveis para alimentar o chargeback/showback.
        """
        entry = FinOpsEntry(
            id=None,
            user_id=user_id,
            module_id=module_id,
            model_name=model_name,
            tokens_input=tokens_input,
            tokens_output=tokens_output,
            cost_estimated=cost_estimated,
            context_tag=context_tag,
            domain=domain,
            product=product,
            agent=agent,
            flow=flow,
            prompt_id=prompt_id,
            integration=integration,
            environment=environment,
            latency_ms=latency_ms,
            storage_bytes=storage_bytes,
        )
        return await self.repo.append(entry)

    async def by_module(self):
        return await self.repo.aggregate_by_module()

    async def by_model(self):
        return await self.repo.aggregate_by_model()

    async def by_day(self, days: int = 7):
        return await self.repo.aggregate_by_day(days)

    async def totals(self):
        return await self.repo.totals()

    async def by_dimension(self, dimension: str) -> list[dict]:
        """Chargeback/showback genérico por dimensão (domain, agent, flow…)."""
        return await self.repo.aggregate_by_dimension(dimension)


# ---------------------------------------------------------------------------
# Orçamentos / Alertas
# ---------------------------------------------------------------------------


@dataclass
class BudgetStatus:
    budget: FinOpsBudget
    spent: float                # custo no período corrente
    remaining: float            # limit - spent (pode ser negativo)
    pct_used: float             # 0..>1 (>1 = estourado)
    severity: str               # 'ok'|'warning'|'critical'


class FinOpsBudgetService:
    """Gerencia orçamentos finops e dispara alertas quando ultrapassados."""

    def __init__(
        self,
        budget_repo: FinOpsBudgetRepository,
        ledger_repo: FinOpsRepository,
    ):
        self.repo = budget_repo
        self.ledger = ledger_repo

    async def list(self) -> list[FinOpsBudget]:
        return await self.repo.list()

    async def get(self, budget_id: UUID) -> FinOpsBudget:
        b = await self.repo.get(budget_id)
        if not b:
            raise ValueError("orçamento não encontrado")
        return b

    async def create(
        self,
        name: str,
        scope_type: str,
        scope_value: str | None,
        period: str,
        limit_brl: float,
        warning_threshold: float = 0.8,
        hard_stop: bool = False,
        notes: str | None = None,
        created_by: UUID | None = None,
    ) -> FinOpsBudget:
        self._validate(name, scope_type, scope_value, period, limit_brl, warning_threshold)
        budget = FinOpsBudget(
            id=new_uuid(),
            name=name.strip(),
            scope_type=BudgetScopeType(scope_type),
            scope_value=(scope_value or None),
            period=BudgetPeriod(period),
            limit_brl=float(limit_brl),
            warning_threshold=float(warning_threshold),
            hard_stop=bool(hard_stop),
            notes=(notes or None),
            created_by=created_by,
        )
        return await self.repo.save(budget)

    async def update(
        self,
        budget_id: UUID,
        name: str | None = None,
        limit_brl: float | None = None,
        warning_threshold: float | None = None,
        hard_stop: bool | None = None,
        notes: str | None = None,
    ) -> FinOpsBudget:
        budget = await self.get(budget_id)
        if name is not None:
            if not name.strip():
                raise ValueError("name não pode ser vazio")
            budget.name = name.strip()
        if limit_brl is not None:
            if limit_brl < 0:
                raise ValueError("limit_brl deve ser >= 0")
            budget.limit_brl = float(limit_brl)
        if warning_threshold is not None:
            if not (0.0 < warning_threshold <= 1.0):
                raise ValueError("warning_threshold deve estar em (0, 1]")
            budget.warning_threshold = float(warning_threshold)
        if hard_stop is not None:
            budget.hard_stop = bool(hard_stop)
        if notes is not None:
            budget.notes = notes or None
        return await self.repo.save(budget)

    async def delete(self, budget_id: UUID) -> None:
        if not await self.repo.delete(budget_id):
            raise ValueError("orçamento não encontrado")

    @staticmethod
    def _validate(name, scope_type, scope_value, period, limit_brl, warning_threshold):
        if not name or not name.strip():
            raise ValueError("name é obrigatório")
        try:
            BudgetScopeType(scope_type)
        except ValueError:
            valid = [s.value for s in BudgetScopeType]
            raise ValueError(f"scope_type inválido. Use {valid}")
        try:
            BudgetPeriod(period)
        except ValueError:
            valid = [p.value for p in BudgetPeriod]
            raise ValueError(f"period inválido. Use {valid}")
        if limit_brl < 0:
            raise ValueError("limit_brl deve ser >= 0")
        if not (0.0 < warning_threshold <= 1.0):
            raise ValueError("warning_threshold deve estar em (0, 1]")
        if scope_type != "global" and not (scope_value and scope_value.strip()):
            raise ValueError(
                f"scope_value é obrigatório quando scope_type='{scope_type}'"
            )

    async def evaluate(self, budget: FinOpsBudget) -> BudgetStatus:
        """Calcula gasto corrente e severidade do orçamento."""
        spent = await self.ledger.current_spend(
            budget.scope_type.value,
            budget.scope_value,
            budget.period.value,
        )
        limit = budget.limit_brl or 0.0
        pct = (spent / limit) if limit > 0 else 0.0
        if pct >= 1.0:
            severity = "critical"
        elif pct >= budget.warning_threshold:
            severity = "warning"
        else:
            severity = "ok"
        return BudgetStatus(
            budget=budget,
            spent=spent,
            remaining=limit - spent,
            pct_used=pct,
            severity=severity,
        )

    async def evaluate_all(self) -> list[BudgetStatus]:
        out = []
        for b in await self.repo.list():
            out.append(await self.evaluate(b))
        # ordena por severidade desc → critical primeiro
        order = {"critical": 0, "warning": 1, "ok": 2}
        out.sort(key=lambda s: (order.get(s.severity, 9), -s.pct_used))
        return out

    async def trigger_alert_if_needed(self, status: BudgetStatus) -> FinOpsAlert | None:
        """Grava alerta no histórico se o status não está OK. Idempotente:
        chamadores podem invocar repetidamente sem duplicar trilha (a
        deduplicação é deixada para uma camada de notificação externa)."""
        if status.severity == "ok":
            return None
        alert = FinOpsAlert(
            id=new_uuid(),
            budget_id=status.budget.id,
            severity=status.severity,
            cost_observed=status.spent,
            limit_reference=status.budget.limit_brl,
        )
        return await self.repo.append_alert(alert)

    async def recent_alerts(self, limit: int = 20) -> list[FinOpsAlert]:
        return await self.repo.list_recent_alerts(limit)

    # --- XLSX template / import ---

    @staticmethod
    def xlsx_template() -> bytes:
        """Gera o template xlsx para upload em massa de orçamentos."""
        return _build_xlsx_template(
            sheet_name="orçamentos",
            headers=_BUDGET_HEADERS,
            examples=_BUDGET_EXAMPLES,
            helper=(
                "Preencha um orçamento por linha. scope_type ∈ "
                "{global, module, model, user, domain, environment, agent}. "
                "Para 'global' deixe scope_value vazio. period ∈ "
                "{daily, weekly, monthly}. warning_threshold em (0..1] (ex.: 0.8). "
                "hard_stop = true|false. Linhas em branco e linhas-exemplo "
                "podem ser apagadas — reimportar é idempotente por NOME apenas "
                "no sentido de criar duplicado, então NÃO reimporte sem revisar."
            ),
        )

    async def import_xlsx(
        self,
        file_bytes: bytes,
        created_by: UUID | None = None,
    ) -> dict[str, Any]:
        """Lê xlsx e cria orçamentos. Retorna ``{imported, errors:[{row,error}]}``.

        Cada linha é validada e gravada de forma independente — uma linha
        inválida não cancela as outras.
        """
        try:
            records = _parse_xlsx_rows(file_bytes, _BUDGET_HEADERS)
        except ValueError as e:
            return {"imported": 0, "errors": [{"row": 0, "error": str(e)}]}

        imported = 0
        errors: list[dict] = []
        for i, rec in enumerate(records, start=1):
            try:
                await self.create(
                    name=str(rec.get("name") or "").strip(),
                    scope_type=str(rec.get("scope_type") or "").strip(),
                    scope_value=(str(rec.get("scope_value")).strip() if rec.get("scope_value") else None),
                    period=str(rec.get("period") or "monthly").strip(),
                    limit_brl=float(rec.get("limit_brl") or 0.0),
                    warning_threshold=float(rec.get("warning_threshold") or 0.8),
                    hard_stop=_coerce_bool(rec.get("hard_stop")),
                    notes=(str(rec.get("notes")).strip() if rec.get("notes") else None),
                    created_by=created_by,
                )
                imported += 1
            except (ValueError, TypeError) as e:
                errors.append({"row": i, "error": str(e)})
        return {"imported": imported, "errors": errors}


# ---------------------------------------------------------------------------
# Políticas de modelo
# ---------------------------------------------------------------------------


class FinOpsPolicyService:
    """CRUD + lookup de políticas de uso de modelo (custo × valor × risco)."""

    def __init__(self, repo: FinOpsModelPolicyRepository):
        self.repo = repo

    async def list(self) -> list[FinOpsModelPolicy]:
        return await self.repo.list()

    async def get_by_model(self, model_name: str) -> FinOpsModelPolicy | None:
        return await self.repo.get_by_model(model_name)

    async def upsert(
        self,
        model_name: str,
        risk_tier: str = "medium",
        value_tier: str = "medium",
        max_cost_per_call: float | None = None,
        max_tokens_per_call: int | None = None,
        allowed_features: list[str] | None = None,
        rationale: str | None = None,
        enabled: bool = True,
    ) -> FinOpsModelPolicy:
        if not model_name or not model_name.strip():
            raise ValueError("model_name é obrigatório")
        try:
            RiskTier(risk_tier)
        except ValueError:
            raise ValueError("risk_tier deve ser low|medium|high")
        try:
            ValueTier(value_tier)
        except ValueError:
            raise ValueError("value_tier deve ser low|medium|high")
        if max_cost_per_call is not None and max_cost_per_call < 0:
            raise ValueError("max_cost_per_call deve ser >= 0")
        existing = await self.repo.get_by_model(model_name.strip())
        policy = FinOpsModelPolicy(
            id=existing.id if existing else new_uuid(),
            model_name=model_name.strip(),
            risk_tier=RiskTier(risk_tier),
            value_tier=ValueTier(value_tier),
            max_cost_per_call=max_cost_per_call,
            max_tokens_per_call=max_tokens_per_call,
            allowed_features=allowed_features,
            rationale=rationale,
            enabled=bool(enabled),
        )
        return await self.repo.save(policy)

    async def delete(self, policy_id: UUID) -> None:
        if not await self.repo.delete(policy_id):
            raise ValueError("política não encontrada")

    @staticmethod
    def xlsx_template() -> bytes:
        """Gera o template xlsx para upload em massa de políticas de modelo."""
        return _build_xlsx_template(
            sheet_name="políticas",
            headers=_POLICY_HEADERS,
            examples=_POLICY_EXAMPLES,
            helper=(
                "Preencha uma política por linha. risk_tier/value_tier ∈ "
                "{low, medium, high}. max_cost_per_call em R$ (vazio = sem cap). "
                "max_tokens_per_call vazio = sem cap. allowed_features = csv "
                "(ex.: 'radar,churn'); vazio = todas as features. enabled = "
                "true|false. Política reimportada para o mesmo model_name faz "
                "UPSERT — atualiza em vez de duplicar."
            ),
        )

    async def import_xlsx(self, file_bytes: bytes) -> dict[str, Any]:
        """Lê xlsx e faz upsert de políticas. Retorna ``{imported, errors}``."""
        try:
            records = _parse_xlsx_rows(file_bytes, _POLICY_HEADERS)
        except ValueError as e:
            return {"imported": 0, "errors": [{"row": 0, "error": str(e)}]}

        imported = 0
        errors: list[dict] = []
        for i, rec in enumerate(records, start=1):
            try:
                features_raw = rec.get("allowed_features")
                features: list[str] | None = None
                if features_raw:
                    if isinstance(features_raw, list):
                        features = [str(x).strip() for x in features_raw if str(x).strip()]
                    else:
                        # csv ou JSON string — aceita os dois
                        s = str(features_raw).strip()
                        if s.startswith("["):
                            try:
                                features = [str(x).strip() for x in _json.loads(s)]
                            except _json.JSONDecodeError:
                                features = [t.strip() for t in s.split(",") if t.strip()]
                        else:
                            features = [t.strip() for t in s.split(",") if t.strip()]
                    if not features:
                        features = None
                await self.upsert(
                    model_name=str(rec.get("model_name") or "").strip(),
                    risk_tier=str(rec.get("risk_tier") or "medium").strip().lower(),
                    value_tier=str(rec.get("value_tier") or "medium").strip().lower(),
                    max_cost_per_call=(float(rec["max_cost_per_call"]) if rec.get("max_cost_per_call") not in (None, "") else None),
                    max_tokens_per_call=(int(rec["max_tokens_per_call"]) if rec.get("max_tokens_per_call") not in (None, "") else None),
                    allowed_features=features,
                    rationale=(str(rec.get("rationale")).strip() if rec.get("rationale") else None),
                    enabled=_coerce_bool(rec.get("enabled")) if rec.get("enabled") is not None else True,
                )
                imported += 1
            except (ValueError, TypeError) as e:
                errors.append({"row": i, "error": str(e)})
        return {"imported": imported, "errors": errors}

    async def is_allowed(
        self,
        model_name: str,
        feature: str | None = None,
        estimated_cost: float | None = None,
    ) -> tuple[bool, str | None]:
        """Decide se uma chamada pode prosseguir sob a política do modelo.

        Retorna ``(True, None)`` se permitida ou ``(False, motivo)`` se
        bloqueada — mensagem é exibível ao usuário final.
        """
        policy = await self.repo.get_by_model(model_name)
        if not policy:
            return True, None  # sem política = permitido (default open)
        if not policy.enabled:
            return False, f"modelo {model_name} desabilitado por política"
        if (
            feature
            and policy.allowed_features
            and feature not in policy.allowed_features
        ):
            return False, (
                f"feature '{feature}' não está na whitelist da política do "
                f"modelo {model_name}"
            )
        if (
            estimated_cost is not None
            and policy.max_cost_per_call is not None
            and estimated_cost > policy.max_cost_per_call
        ):
            return False, (
                f"custo estimado R$ {estimated_cost:.4f} excede o cap da "
                f"política do modelo {model_name} (R$ {policy.max_cost_per_call:.4f})"
            )
        return True, None


# ---------------------------------------------------------------------------
# Cost-aware routing
# ---------------------------------------------------------------------------


class CostAwareRouter:
    """Recomenda o modelo a usar dado um conjunto de candidatos, custos
    estimados e o estado vigente das políticas + orçamentos.

    Estratégia (em ordem):

      1. Filtra modelos cuja política bloqueia a feature ou supera cap.
      2. Filtra modelos cujos orçamentos estouraram (com hard_stop=True).
      3. Entre os restantes, escolhe o de **menor custo estimado** que ainda
         atenda ao requisito de qualidade (``min_value_tier``).
      4. Em caso de empate, prefere o de menor risco.
    """

    def __init__(
        self,
        policy_svc: FinOpsPolicyService,
        budget_svc: FinOpsBudgetService,
    ):
        self.policy_svc = policy_svc
        self.budget_svc = budget_svc

    async def recommend(
        self,
        candidates: list[dict[str, Any]],
        feature: str | None = None,
        min_value_tier: str = "low",
    ) -> dict[str, Any]:
        """Cada candidato é ``{"model": str, "estimated_cost": float}``.

        Retorna ``{model, reason, blocked: [{model, reason}]}``.
        """
        min_rank = {"low": 0, "medium": 1, "high": 2}[min_value_tier]
        survivors: list[tuple[dict, FinOpsModelPolicy | None]] = []
        blocked: list[dict] = []

        for c in candidates:
            policy = await self.policy_svc.get_by_model(c["model"])
            ok, reason = await self.policy_svc.is_allowed(
                c["model"], feature=feature, estimated_cost=c.get("estimated_cost"),
            )
            if not ok:
                blocked.append({"model": c["model"], "reason": reason})
                continue
            if policy and min_rank > {"low": 0, "medium": 1, "high": 2}[policy.value_tier.value]:
                blocked.append({
                    "model": c["model"],
                    "reason": f"value_tier ({policy.value_tier.value}) abaixo do exigido ({min_value_tier})",
                })
                continue
            # checa orçamentos com hard_stop estourados que cobrem este modelo
            if await self._is_hard_stopped(c["model"]):
                blocked.append({
                    "model": c["model"],
                    "reason": "orçamento com hard_stop estourado",
                })
                continue
            survivors.append((c, policy))

        if not survivors:
            return {
                "model": None,
                "reason": "todos os candidatos bloqueados por política/orçamento",
                "blocked": blocked,
            }

        # menor custo, depois menor risco
        risk_rank = {"low": 0, "medium": 1, "high": 2}
        survivors.sort(key=lambda t: (
            t[0].get("estimated_cost", 0.0),
            risk_rank.get(t[1].risk_tier.value if t[1] else "medium", 1),
        ))
        chosen = survivors[0][0]
        return {
            "model": chosen["model"],
            "reason": (
                f"menor custo estimado (R$ {chosen.get('estimated_cost', 0):.4f}) "
                f"entre {len(survivors)} modelo(s) elegível(is)"
            ),
            "blocked": blocked,
        }

    async def _is_hard_stopped(self, model_name: str) -> bool:
        """True se há algum orçamento de escopo `model=<this>` ou `global`
        com hard_stop=True que esteja estourado."""
        for status in await self.budget_svc.evaluate_all():
            b = status.budget
            applies = (
                (b.scope_type == BudgetScopeType.global_)
                or (b.scope_type == BudgetScopeType.model and b.scope_value == model_name)
            )
            if applies and b.hard_stop and status.severity == "critical":
                return True
        return False
